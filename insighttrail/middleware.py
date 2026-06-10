from flask import request, g, jsonify, render_template, Blueprint, send_file
import time
import os
import json
import threading
import glob
from datetime import datetime, timedelta, timezone
from io import BytesIO
from collections import deque
from .logger import setup_logger, log_request, log_error, get_logger_stats, should_log_success
from .metrics import record_metrics, get_metrics
from .traces import trace_request
import importlib.metadata
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class FlaskInsightTrail:
    def __init__(self, app, log_file=None, log_level='INFO', max_file_size=1 * 1024 * 1024, backup_count=5,
                 enable_ui=True, url_prefix='/insight', capture_runtime=False,
                 capture_system_metrics=False, capture_env_vars=False, env_allowlist=None,
                 dependency_check=None, ultra_light_mode=False, enable_charts=None,
                 ui_refresh_seconds=10, track_internal_requests=False,
                 async_logging=True, log_queue_size=5000,
                 success_log_sample_rate=1.0, slow_request_threshold_ms=None,
                 dependency_cache_ttl_seconds=21600, dependency_async_refresh=True,
                 dependency_request_timeout=2, enable_excel_reports=True,
                 report_max_rows=200000, report_timezone='UTC'):
        """
        Initialize InsightTrail middleware.

        Args:
            app: Flask application instance
            log_file: Path to log file. Defaults to 'insighttrail.log' in the parent directory of the app's root path.
            log_level: The logging level to use, e.g., 'INFO', 'DEBUG'.
            max_file_size: Maximum size of log file before rotation
            backup_count: Number of backup files to keep
            enable_ui: Whether to enable the web UI (default: True)
            url_prefix: URL prefix for InsightTrail routes (default: /insight)
        """
        self.app = app
        self.capture_runtime = capture_runtime
        self.capture_system_metrics = capture_system_metrics
        self.capture_env_vars = capture_env_vars
        self.env_allowlist = env_allowlist or []
        self.url_prefix = '/' + url_prefix.strip('/') if url_prefix else '/insight'
        self.track_internal_requests = track_internal_requests
        self.async_logging = async_logging
        self.log_queue_size = max(100, int(log_queue_size))
        self.success_log_sample_rate = max(0.0, min(1.0, float(success_log_sample_rate)))
        self.slow_request_threshold_ms = float(slow_request_threshold_ms) if slow_request_threshold_ms is not None else None
        self.dependency_cache_ttl_seconds = max(60, int(dependency_cache_ttl_seconds))
        self.dependency_async_refresh = dependency_async_refresh
        self.dependency_request_timeout = max(1, int(dependency_request_timeout))
        self.enable_excel_reports = enable_excel_reports
        self.report_max_rows = max(1000, int(report_max_rows))
        self.report_timezone = report_timezone
        self.ultra_light_mode = ultra_light_mode
        self.dependency_check = (not ultra_light_mode) if dependency_check is None else dependency_check
        self.enable_charts = (not ultra_light_mode) if enable_charts is None else enable_charts
        self.ui_refresh_seconds = max(2, int(ui_refresh_seconds))
        self.required_packages = self._load_required_packages(app.root_path)
        self._log_cache = deque(maxlen=3000)
        self._log_file_offset = 0
        self._next_log_id = 1
        self._dependency_cache = {}
        self._dependency_refresh_in_progress = False
        
        if log_file is None:
            # Default to a 'logs' directory in the parent of the app's root path
            app_parent_dir = os.path.dirname(app.root_path)
            log_file = os.path.join(app_parent_dir, 'logs', 'insighttrail.log')

        setup_logger(
            log_file,
            log_level,
            max_file_size,
            backup_count,
            async_logging=self.async_logging,
            log_queue_size=self.log_queue_size,
        )
        self.log_file = log_file
        self._init_app(app)

        if enable_ui:
            self._setup_ui(self.url_prefix)

    def _load_required_packages(self, start_path):
        """
        Traverse up from start_path to find and parse a requirements.txt file.
        This helps identify the host application's dependencies.
        """
        current_path = start_path
        # Limit search to 5 levels up to avoid scanning the whole filesystem
        for _ in range(5):
            requirements_file = os.path.join(current_path, 'requirements.txt')
            if os.path.exists(requirements_file):
                try:
                    with open(requirements_file, 'r') as f:
                        packages = []
                        for line in f:
                            line = line.strip()
                            if line and not line.startswith('#'):
                                # Basic parsing: remove version specifiers and comments
                                package_name = line.split('#')[0].strip()
                                package_name = package_name.split('==')[0].split('>=')[0].split('<=')[0].split('~=')[0].split('<')[0].split('>')[0].split('!=')[0].strip()
                                if package_name:
                                    packages.append(package_name.lower())
                        return packages
                except IOError:
                    return []  # Return empty list on read error
            
            parent = os.path.dirname(current_path)
            if parent == current_path:  # Reached the filesystem root
                break
            current_path = parent
            
        return []  # Return empty list if no requirements.txt is found

    def _get_package_info(self):
        """
        Gathers information about installed packages, highlighting those required
        by the host application and InsightTrail itself.
        """
        packages = []
        # Combine app's requirements with InsightTrail's key dependencies for highlighting
        insighttrail_deps = {'flask', 'waitress', 'psutil', 'requests'}
        app_deps = set(self.required_packages)
        required_set = app_deps.union(insighttrail_deps)

        stale_keys = []

        for dist in importlib.metadata.distributions():
            try:
                name = dist.metadata['Name']
                version = dist.version
                key = name.lower()
                is_prerelease = any(tag in version.lower() for tag in ('a', 'b', 'rc', 'dev', 'alpha', 'beta'))
                package = {
                    'name': key,
                    'current_version': version,
                    'latest_version': version,
                    'required': key in required_set,
                    'description': dist.metadata.get('Summary'),
                    'stability': 'pre-release' if is_prerelease else 'stable'
                }

                if self.dependency_check:
                    cache_data, is_fresh = self._get_cached_dependency_info(dist.key)
                    if cache_data is not None:
                        package['latest_version'] = cache_data.get('latest_version', package['latest_version'])
                        if not package['description']:
                            package['description'] = cache_data.get('description')
                        package['stability'] = cache_data.get('stability', package['stability'])
                    if not is_fresh:
                        stale_keys.append(key)

                packages.append(package)
            except Exception:
                continue

        if self.dependency_check and stale_keys and self.dependency_async_refresh:
            self._refresh_dependency_cache_background(stale_keys)

        # Sort packages: required first, then alphabetically
        return sorted(packages, key=lambda x: (not x['required'], x['name'].lower()))

    def _get_cached_dependency_info(self, package_name):
        now = time.time()
        entry = self._dependency_cache.get(package_name)
        if entry and (now - entry.get('fetched_at', 0) <= self.dependency_cache_ttl_seconds):
            return entry, True

        if not self.dependency_async_refresh:
            fresh = self._fetch_dependency_info(package_name)
            if fresh is not None:
                self._dependency_cache[package_name] = fresh
                return fresh, True

        return entry, False

    def _fetch_dependency_info(self, package_name):
        try:
            pypi_url = f"https://pypi.org/pypi/{package_name}/json"
            response = requests.get(pypi_url, timeout=self.dependency_request_timeout)
            if response.status_code != 200:
                return None
            pypi_data = response.json()
            latest_version = pypi_data['info']['version']
            summary = pypi_data['info'].get('summary')
            latest_lower = str(latest_version).lower()
            stability = 'pre-release' if any(tag in latest_lower for tag in ('a', 'b', 'rc', 'dev', 'alpha', 'beta')) else 'stable'
            return {
                'latest_version': latest_version,
                'description': summary,
                'stability': stability,
                'fetched_at': time.time(),
            }
        except (requests.RequestException, KeyError, ValueError):
            return None

    def _refresh_dependency_cache_background(self, package_names):
        if self._dependency_refresh_in_progress:
            return

        unique_names = sorted(set(package_names))

        def _worker():
            self._dependency_refresh_in_progress = True
            try:
                for package_name in unique_names:
                    data = self._fetch_dependency_info(package_name)
                    if data is not None:
                        self._dependency_cache[package_name] = data
            finally:
                self._dependency_refresh_in_progress = False

        thread = threading.Thread(target=_worker, name='insighttrail-dependency-refresh')
        thread.daemon = True
        thread.start()

    def _init_app(self, app):
        @app.before_request
        def before_request():
            g.start_time = time.time()
            trace_request(request)

        @app.after_request
        def after_request(response):
            if not self.track_internal_requests and request.path.startswith(self.url_prefix):
                return response
            duration = time.time() - g.start_time
            record_metrics(request, response, duration)

            is_error_status = response.status_code >= 400
            should_log = is_error_status or should_log_success(
                duration,
                success_log_sample_rate=self.success_log_sample_rate,
                slow_request_threshold_ms=self.slow_request_threshold_ms,
            )
            if should_log:
                log_request(
                    request,
                    response,
                    duration,
                    capture_runtime=self.capture_runtime,
                    capture_system_metrics=self.capture_system_metrics,
                    capture_env_vars=self.capture_env_vars,
                    env_allowlist=self.env_allowlist,
                )
            return response

        @app.teardown_request
        def teardown_request(exception=None):
            if exception is not None:
                if not self.track_internal_requests and request.path.startswith(self.url_prefix):
                    return
                duration = time.time() - g.start_time
                log_error(
                    request,
                    exception,
                    duration,
                    capture_runtime=self.capture_runtime,
                    capture_system_metrics=self.capture_system_metrics,
                    capture_env_vars=self.capture_env_vars,
                    env_allowlist=self.env_allowlist,
                )

    def _parse_log_file(self):
        self._refresh_log_cache()
        return list(self._log_cache)

    def _refresh_log_cache(self):
        try:
            if not os.path.exists(self.log_file):
                return

            current_size = os.path.getsize(self.log_file)
            if current_size < self._log_file_offset:
                self._log_file_offset = 0
                self._log_cache.clear()
                self._next_log_id = 1

            with open(self.log_file, 'r', encoding='utf-8', errors='replace') as f:
                f.seek(self._log_file_offset)
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        log_entry = json.loads(line)
                        log_entry['_id'] = self._next_log_id
                        self._next_log_id += 1
                        self._log_cache.append(log_entry)
                    except (json.JSONDecodeError, ValueError):
                        continue
                self._log_file_offset = f.tell()
        except Exception as e:
            print(f"Error reading log file: {e}")

    def _get_logs_page(self, limit=100, cursor=None):
        self._refresh_log_cache()
        max_limit = 500
        safe_limit = max(1, min(int(limit), max_limit))
        logs = list(self._log_cache)

        if cursor is not None:
            cursor_id = int(cursor)
            filtered = [log for log in logs if int(log.get('_id', 0)) > cursor_id]
            page = filtered[:safe_limit]
        else:
            page = logs[-safe_limit:]

        next_cursor = page[-1]['_id'] if page else (cursor if cursor is not None else 0)
        return {
            'logs': page,
            'cursor': next_cursor,
            'has_more': len(logs) > len(page) if cursor is None else False,
        }

    def _setup_ui(self, normalized_prefix):
        # Create a blueprint for InsightTrail UI
        insight_bp = Blueprint('insighttrail', __name__,
                               template_folder='templates',
                               static_folder='static',
                               url_prefix=normalized_prefix)

        @insight_bp.route('/', strict_slashes=False)
        def index():
            return render_template(
                "insighttrail_dashboard.html",
                insight_base_url=normalized_prefix,
                ui_refresh_seconds=self.ui_refresh_seconds,
                enable_charts=self.enable_charts,
                dependency_check=self.dependency_check,
            )

        @insight_bp.route('/api/packages')
        def get_packages():
            return jsonify(self._get_package_info())

        @insight_bp.route('/api/logs')
        def get_logs():
            try:
                limit = request.args.get('limit', default=100, type=int)
                cursor = request.args.get('cursor', default=None, type=int)
                page = self._get_logs_page(limit=limit, cursor=cursor)
                return jsonify(page)
            except Exception as e:
                print(f"Error in get_logs: {e}")
                return jsonify({"error": str(e)}), 500

        @insight_bp.route('/api/analytics/logs', methods=['GET'])
        def fetch_logs():
            try:
                limit = request.args.get('limit', default=100, type=int)
                cursor = request.args.get('cursor', default=None, type=int)
                page = self._get_logs_page(limit=limit, cursor=cursor)
                metrics = get_metrics()
                return jsonify({
                    'logs': page['logs'],
                    'cursor': page['cursor'],
                    'has_more': page['has_more'],
                    'metrics': metrics,
                    'logger': get_logger_stats(),
                })
            except Exception as e:
                print(f"Error in fetch_logs: {e}")
                return jsonify({"error": str(e)}), 500

        @insight_bp.route('/api/analytics/search', methods=['GET'])
        def search_by_trace_id():
            try:
                trace_id = request.args.get('trace_id')
                self._refresh_log_cache()
                logs = list(self._log_cache)
                result = [log for log in logs if log.get("trace_id") == trace_id]
                metrics = get_metrics()
                return jsonify({
                    'logs': result,
                    'metrics': metrics,
                    'logger': get_logger_stats(),
                })
            except Exception as e:
                print(f"Error in search_by_trace_id: {e}")
                return jsonify({"error": str(e)}), 500

        @insight_bp.route('/api/reports/excel', methods=['GET'])
        def export_excel_report():
            if not self.enable_excel_reports:
                return jsonify({'error': 'Excel report export is disabled'}), 403
            try:
                start_dt, end_dt = self._resolve_report_range(request.args)
                report_rows = self._collect_logs_for_range(start_dt, end_dt)
                include_sheets = self._resolve_report_sheets(request.args)
                workbook_bytes = self._build_excel_report(report_rows, start_dt, end_dt, include_sheets)
                file_name = f"insighttrail_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
                return send_file(
                    workbook_bytes,
                    as_attachment=True,
                    download_name=file_name,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
            except ValueError as e:
                return jsonify({'error': str(e)}), 400
            except Exception as e:
                return jsonify({'error': f'Failed to generate report: {e}'}), 500

        @insight_bp.route('/api/reports/estimate', methods=['GET'])
        def estimate_excel_report_rows():
            if not self.enable_excel_reports:
                return jsonify({'error': 'Excel report export is disabled'}), 403
            try:
                start_dt, end_dt = self._resolve_report_range(request.args)
                estimated_rows = self._estimate_logs_for_range(start_dt, end_dt)
                return jsonify({
                    'estimated_rows': estimated_rows,
                    'report_max_rows': self.report_max_rows,
                    'truncated': estimated_rows >= self.report_max_rows,
                })
            except ValueError as e:
                return jsonify({'error': str(e)}), 400
            except Exception as e:
                return jsonify({'error': f'Failed to estimate rows: {e}'}), 500

        self.app.register_blueprint(insight_bp)

    def _resolve_report_sheets(self, args):
        include_raw = args.get('include', '')
        if not include_raw:
            return {'summary', 'requests', 'errors', 'dependencies'}
        allowed = {'summary', 'requests', 'errors', 'dependencies'}
        selected = {item.strip().lower() for item in include_raw.split(',') if item.strip()}
        selected = selected.intersection(allowed)
        if not selected:
            return {'summary', 'requests', 'errors', 'dependencies'}
        return selected

        # Register the blueprint with the main app
        self.app.register_blueprint(insight_bp)

    def _local_tz(self):
        return datetime.now().astimezone().tzinfo

    def _resolve_report_range(self, args):
        preset = (args.get('preset') or '').strip().lower()
        now = datetime.now(self._local_tz())
        if preset:
            mapping = {'1d': 1, '7d': 7, '1m': 30, '6m': 180}
            if preset not in mapping:
                raise ValueError('Invalid preset. Use 1d, 7d, 1m, or 6m.')
            return now - timedelta(days=mapping[preset]), now

        start_raw = args.get('start')
        end_raw = args.get('end')
        if not start_raw or not end_raw:
            raise ValueError('Provide preset or both start and end in ISO format.')

        start_dt = self._parse_utc_iso(start_raw)
        end_dt = self._parse_utc_iso(end_raw)
        if start_dt >= end_dt:
            raise ValueError('start must be earlier than end.')
        return start_dt, end_dt

    def _parse_utc_iso(self, value):
        normalized = value.strip().replace('Z', '+00:00')
        dt = datetime.fromisoformat(normalized)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=self._local_tz())
        return dt.astimezone(self._local_tz())

    def _log_files(self):
        log_dir = os.path.dirname(self.log_file) or '.'
        base_name = os.path.basename(self.log_file)
        files = glob.glob(os.path.join(log_dir, f"{base_name}*"))
        return sorted([f for f in files if os.path.isfile(f)], key=os.path.getmtime)

    def _collect_logs_for_range(self, start_dt, end_dt):
        rows = []
        for file_path in self._log_files():
            try:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as handle:
                    for line in handle:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            entry = json.loads(line)
                            ts_raw = entry.get('timestamp')
                            if not ts_raw:
                                continue
                            ts = self._parse_utc_iso(ts_raw)
                            if start_dt <= ts <= end_dt:
                                entry['_parsed_ts'] = ts
                                rows.append(entry)
                                if len(rows) >= self.report_max_rows:
                                    break
                        except Exception:
                            continue
                if len(rows) >= self.report_max_rows:
                    break
            except IOError:
                continue

        rows.sort(key=lambda r: r.get('_parsed_ts', datetime.min.replace(tzinfo=timezone.utc)))
        return rows

    def _estimate_logs_for_range(self, start_dt, end_dt):
        count = 0
        for file_path in self._log_files():
            try:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as handle:
                    for line in handle:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            entry = json.loads(line)
                            ts_raw = entry.get('timestamp')
                            if not ts_raw:
                                continue
                            ts = self._parse_utc_iso(ts_raw)
                            if start_dt <= ts <= end_dt:
                                count += 1
                                if count >= self.report_max_rows:
                                    return count
                        except Exception:
                            continue
            except IOError:
                continue
        return count

    def _build_excel_report(self, rows, start_dt, end_dt, include_sheets):
        workbook = Workbook()
        summary = workbook.active
        summary.title = 'Summary'

        total = len(rows)
        errors = len([r for r in rows if (r.get('request') or {}).get('status', 0) >= 400 or r.get('error')])
        latencies = [float((r.get('request') or {}).get('duration_ms') or 0) for r in rows]
        avg_latency = (sum(latencies) / len(latencies)) if latencies else 0
        p95_latency = sorted(latencies)[int(0.95 * (len(latencies) - 1))] if latencies else 0

        summary_rows = [
            ('InsightTrail Report', ''),
            ('Generated At', datetime.now(self._local_tz()).isoformat()),
            ('Range Start', start_dt.isoformat()),
            ('Range End', end_dt.isoformat()),
            ('Total Requests', total),
            ('Error Count', errors),
            ('Error Rate %', round((errors / total) * 100, 2) if total else 0),
            ('Average Latency (ms)', round(avg_latency, 2)),
            ('P95 Latency (ms)', round(p95_latency, 2)),
        ]
        for idx, (k, v) in enumerate(summary_rows, start=1):
            summary.cell(row=idx, column=1, value=k)
            summary.cell(row=idx, column=2, value=v)
        summary['A1'].font = Font(bold=True, size=14)

        if 'summary' not in include_sheets:
            workbook.remove(summary)

        requests_sheet = None
        if 'requests' in include_sheets:
            requests_sheet = workbook.create_sheet('Requests')
            request_headers = ['timestamp', 'trace_id', 'method', 'path', 'status', 'duration_ms', 'client']
            requests_sheet.append(request_headers)
            for row in rows:
                req = row.get('request') or {}
                requests_sheet.append([
                    row.get('timestamp'), row.get('trace_id'), req.get('method'), req.get('path'),
                    req.get('status'), req.get('duration_ms'), req.get('client')
                ])

        error_sheet = None
        if 'errors' in include_sheets:
            error_sheet = workbook.create_sheet('Errors')
            error_headers = ['timestamp', 'trace_id', 'method', 'path', 'status', 'error_type', 'error_message']
            error_sheet.append(error_headers)
            for row in rows:
                req = row.get('request') or {}
                err_raw = row.get('error')
                if isinstance(err_raw, dict):
                    err = err_raw
                elif isinstance(err_raw, str):
                    err = {'message': err_raw, 'type': 'Error'}
                else:
                    err = {}
                status = req.get('status') or 0
                if status >= 400 or err:
                    error_message = err.get('message') or err.get('traceback') or f"HTTP {status}" if status >= 400 else ''
                    error_sheet.append([
                        row.get('timestamp'), row.get('trace_id'), req.get('method'), req.get('path'),
                        status, err.get('type') or ('HTTPError' if status >= 400 else ''), error_message
                    ])

        dep_sheet = None
        if 'dependencies' in include_sheets:
            dep_sheet = workbook.create_sheet('Dependencies')
            dep_headers = ['name', 'current_version', 'latest_version', 'stability', 'required', 'description']
            dep_sheet.append(dep_headers)
            for dep in self._get_package_info()[:500]:
                dep_sheet.append([
                    dep.get('name'), dep.get('current_version'), dep.get('latest_version'),
                    dep.get('stability'), dep.get('required'), dep.get('description')
                ])

        header_fill = PatternFill(start_color='146C94', end_color='146C94', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for ws in (requests_sheet, error_sheet, dep_sheet):
            if ws is None:
                continue
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

        if not workbook.worksheets:
            fallback = workbook.create_sheet('Summary')
            fallback.append(['InsightTrail Report'])

        for ws in workbook.worksheets:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 60)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

